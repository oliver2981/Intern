import os
import json
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def set_excel_style(ws):
    # 样式设置
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style="thin"),
                         right=Side(style="thin"),
                         top=Side(style="thin"),
                         bottom=Side(style="thin"))

    data_font = Font(name="微软雅黑", size=10)
    data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # 应用标题样式
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # 应用数据样式
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

    # 智能列宽控制
    max_widths = {
        "步骤": 40,  # 限制最大40字符
        "预期": 40,
        "用例标题": 25,
        "所属模块": 20,
        "前置条件": 30
    }

    for col in ws.columns:
        col_letter = col[0].column_letter
        header = ws[f"{col_letter}1"].value

        if header in max_widths:
            # 计算列宽（不超过最大值）
            max_content_len = max(
                len(str(cell.value)) if cell.value else 0
                for cell in col
            )
            final_width = min(max_content_len, max_widths[header]) * 1.2
            ws.column_dimensions[col_letter].width = final_width
        else:
            # 常规列宽计算
            max_length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in col
            )
            ws.column_dimensions[col_letter].width = (max_length + 2) * 1.2

    # 冻结标题行
    ws.freeze_panes = "A2"


def process_json_files():
    input_dir = r"D:\test"
    output_dir = r"D:\test_output"

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    all_data = []
    chapter_data = {}
    columns = ["所属模块", "用例标题", "前置条件", "步骤", "预期", "优先级", "测试结果", "是否可用"]

    # 第一阶段：数据收集
    for filename in os.listdir(input_dir):
        if filename.endswith(".json"):
            module_name = os.path.splitext(filename)[0]
            match = re.search(r"(\d+)", filename)
            chapter_num = int(match.group(1)) if match else 999

            filepath = os.path.join(input_dir, filename)
            with open(filepath, "r", encoding="utf-8") as f:
                try:
                    data = json.load(f)
                except json.JSONDecodeError:
                    continue

                for case in data:
                    # 生成带编号的步骤和预期
                    steps_list = [f"{i + 1}. {step['step']}" for i, step in enumerate(case["steps"])]
                    expect_list = [f"{i + 1}. {step['expect']}" for i, step in enumerate(case["steps"])]

                    row = {
                        "所属模块": module_name,
                        "用例标题": case["casetitle"],
                        "前置条件": case["precondition"],
                        "步骤": "\n".join(steps_list),
                        "预期": "\n".join(expect_list),
                        "优先级": case["priority"],
                        "测试结果": "",
                        "是否可用": ""
                    }

                    # 存储到章节字典
                    if chapter_num not in chapter_data:
                        chapter_data[chapter_num] = []
                    chapter_data[chapter_num].append(row)
                    all_data.append(row)

    # 第二阶段：生成分章文件
    for ch_num, data in chapter_data.items():
        if ch_num == 999:
            output_filename = "未分类.xlsx"
        else:
            output_filename = f"第{ch_num}章.xlsx"

        chapter_filepath = os.path.join(output_dir, output_filename)
        df = pd.DataFrame(data)
        df.to_excel(chapter_filepath, index=False, columns=columns)

        # 应用样式
        wb = load_workbook(chapter_filepath)
        ws = wb.active
        set_excel_style(ws)
        wb.save(chapter_filepath)

    # 第三阶段：生成排序后的汇总文件
    if chapter_data:
        summary_path = os.path.join(output_dir, "汇总.xlsx")

        # 按章节数字排序（1→2→3...→未分类）
        sorted_chapters = sorted([k for k in chapter_data.keys() if k != 999])
        sorted_data = []

        for ch in sorted_chapters:
            sorted_data.extend(chapter_data[ch])
        if 999 in chapter_data:
            sorted_data.extend(chapter_data[999])

        # 生成并保存
        df_all = pd.DataFrame(sorted_data)
        df_all.to_excel(summary_path, index=False, columns=columns)

        # 应用样式
        wb = load_workbook(summary_path)
        ws = wb.active
        set_excel_style(ws)
        wb.save(summary_path)


if __name__ == "__main__":
    process_json_files()